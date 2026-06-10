#!/usr/bin/env python3
"""Export borderline student assignment table to Excel format."""

from __future__ import annotations

import argparse
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def read_markdown_table(file_path: Path) -> tuple[list[str], list[list[str]]]:
    """Read Markdown table from file."""
    text = file_path.read_text(encoding="utf-8")
    
    # Find table lines
    lines = text.splitlines()
    table_start = -1
    table_end = -1
    for i, line in enumerate(lines):
        if line.startswith("|") and "责任教师" in line:
            table_start = i
        elif table_start > -1 and line.startswith("|---"):
            table_end = i
            break
        elif table_start > -1 and not line.startswith("|"):
            table_end = i
            break
    
    if table_start == -1:
        raise ValueError("Could not find table in Markdown file")
    
    # Extract table data
    table_lines = lines[table_start:table_end]
    
    # Parse header - simpler approach: remove leading/trailing | and split
    header_line = table_lines[0].strip()
    if header_line.startswith("|"):
        header_line = header_line[1:]
    if header_line.endswith("|"):
        header_line = header_line[:-1]
    headers = [h.strip() for h in header_line.split("|")]
    
    # Parse data rows
    data: list[list[str]] = []
    for line in table_lines[2:]:  # Skip header and separator
        if line.startswith("|"):
            # Remove leading/trailing | and split by |
            row_line = line.strip()
            if row_line.startswith("|"):
                row_line = row_line[1:]
            if row_line.endswith("|"):
                row_line = row_line[:-1]
            row = [cell.strip() for cell in row_line.split("|")]
            data.append(row)
    
    print(f"Parsed {len(headers)} headers: {headers}")
    print(f"Parsed {len(data)} data rows")
    if data:
        print(f"First row has {len(data[0])} columns: {data[0]}")
    
    return headers, data


def create_excel(headers: list[str], data: list[list[str]], output_file: Path) -> None:
    """Create Excel file with formatted table."""
    wb = Workbook()
    ws = wb.active
    ws.title = "临界生分工表"
    
    # Column widths
    column_widths = {
        "责任教师": 15,
        "任教班级": 15,
        "临界姓名": 12,
        "临界生信息<br>（本周期待提高内容）": 45,
        "临界生信息（本周期待提高内容）": 45,
        "帮扶核心方向<br>（可多选，在对应□内打√）": 35,
        "帮扶核心方向（可多选，在对应□内打√）": 35,
        "完成时限": 15,
        "进步或待改进点": 45,
        "确认方式": 20,
    }
    
    # Set column widths
    for col_idx, header in enumerate(headers, start=1):
        width = column_widths.get(header, 20)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # Styles
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    header_font = Font(bold=True, size=11)
    cell_font = Font(size=10)
    title_font = Font(bold=True, size=16)
    
    header_alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
    cell_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    title_alignment = Alignment(horizontal='center', vertical='center')
    
    # Write title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value="临界生分工表")
    title_cell.font = title_font
    title_cell.alignment = title_alignment
    
    # Write header
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header.replace("<br>", "\n"))
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
        cell.fill = header_fill
    
    # Write data
    for row_idx, row_data in enumerate(data, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = cell_font
            cell.alignment = cell_alignment
            cell.border = thin_border
    
    # Set row heights
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 45
    for i in range(3, len(data) + 3):
        ws.row_dimensions[i].height = 60
    
    # Page setup
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.leftMargin = 0.7
    ws.page_setup.rightMargin = 0.7
    ws.page_setup.topMargin = 0.75
    ws.page_setup.bottomMargin = 0.75
    ws.page_setup.horizontalCentered = True
    
    wb.save(output_file)


def main() -> int:
    """Main function."""
    parser = argparse.ArgumentParser(description="Export borderline student table to Excel.")
    parser.add_argument("file", type=Path, help="Input Markdown file")
    parser.add_argument("--output", "-o", type=Path, help="Output Excel file (optional)")
    args = parser.parse_args()
    
    try:
        headers, data = read_markdown_table(args.file)
        
        # Determine output file
        if args.output:
            output_file = args.output
        else:
            output_file = args.file.with_suffix(".xlsx")
        
        create_excel(headers, data, output_file)
        
        print(f"Excel file generated successfully: {output_file}")
        return 0
        
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    import sys
    sys.exit(main())
